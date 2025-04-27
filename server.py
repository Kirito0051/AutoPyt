from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import zipfile
from typing import List
import pandas as pd
import io
import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Constants
OUTPUT_DIR = "processed"
REQUIRED_COLUMNS = [
    "Case ID", "Mother name", "User name", "User Phone", "User Role",
    "Child ID", "Child Name", "Child DOB", "Child Weight Zscore",
    "Last Weight Zscore", "Mother Location", "User Block/Project/Tehsil",
    "User Facility/Center"
]

# Ensure output directory exists
os.makedirs(OUTPUT_DIR, exist_ok=True)

def format_excel_worksheet(ws):
    """Format Excel worksheet with table and column widths"""
    try:
        table = Table(displayName=f"Table_{ws.title}", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        for col_idx, col in enumerate(ws.columns, 1):
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    except Exception as e:
        logger.error(f"Error formatting worksheet {ws.title}: {str(e)}")
        raise

def process_dataframe(df: pd.DataFrame) -> tuple:
    """Process dataframe and return segregated data"""
    zscore_column = 'Last Weight Zscore'
    return (
        df[df[zscore_column] <= -3],  # SUW
        df[(df[zscore_column] > -3) & (df[zscore_column] <= -2)],  # MUW
        df[(df[zscore_column] > -2) & (df[zscore_column] <= -1)]   # Mild
    )

@app.post("/upload")
async def upload_files(files: List[UploadFile] = File(...)):
    try:
        if not files:
            raise HTTPException(status_code=400, detail="No files provided")

        processed_files = []

        for file in files:
            logger.info(f"Processing file: {file.filename}")
            
            # Create processed filename using original name
            processed_filename = f"processed_{file.filename}"
            
            # Read and validate CSV
            contents = await file.read()
            try:
                df = pd.read_csv(io.BytesIO(contents), low_memory=False, on_bad_lines='skip')
                logger.info(f"Original rows in {file.filename}: {len(df)}")
                
                # Validate required columns
                missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
                if missing_columns:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Missing columns in {file.filename}: {missing_columns}"
                    )

                # Filter and process data
                df = df[REQUIRED_COLUMNS]
                df_suw, df_muw, df_mild = process_dataframe(df)

                # Create Excel file
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for name, data in [("SUW", df_suw), ("MUW", df_muw), ("Mild", df_mild)]:
                        data.to_excel(writer, sheet_name=name, index=False)
                        
                output.seek(0)
                wb = load_workbook(filename=output)
                
                # Format each worksheet
                for sheet_name in wb.sheetnames:
                    format_excel_worksheet(wb[sheet_name])

                # Save processed file
                output_path = os.path.join(OUTPUT_DIR, processed_filename)
                wb.save(output_path)
                
                processed_files.append(output_path)
                logger.info(f"Successfully processed {file.filename}")

            except Exception as e:
                logger.error(f"Error processing {file.filename}: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error processing {file.filename}: {str(e)}")

        # Create a zip of all processed files
        if processed_files:
            zip_filename = f"processed_files_{datetime.now().strftime('%Y%m%d%H%M%S')}.zip"
            zip_filepath = os.path.join(OUTPUT_DIR, zip_filename)
            
            with zipfile.ZipFile(zip_filepath, 'w') as zipf:
                for file in processed_files:
                    zipf.write(file, os.path.basename(file))

            return FileResponse(
                zip_filepath,
                media_type="application/zip",
                filename=zip_filename
            )
        
        raise HTTPException(status_code=500, detail="No files were processed successfully")

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy"}
